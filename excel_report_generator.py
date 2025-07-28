"""
AIRISS v4.0 Excel Report Generator
OK Financial Group Standard Excel Output
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.axis import DateAxis
from typing import List, Dict, Any
from pathlib import Path
import json

from airiss_analyzer_v4 import EmployeeAnalysis, OrganizationAnalysis, AIRISSCompetency

class ExcelReportGenerator:
    """AIRISS v4.0 Excel 리포트 생성기"""
    
    def __init__(self):
        # 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 등급별 색상
        self.grade_colors = {
            'S': 'FFD700',      # Gold
            'A+': 'FFA500',     # Orange
            'A': '90EE90',      # Light Green
            'B+': '87CEEB',     # Sky Blue
            'B': 'DDA0DD',      # Plum
            'C': 'F0E68C',      # Khaki
            'D': 'FFB6C1'       # Light Pink
        }
    
    def generate_report(self, 
                       individual_analyses: List[EmployeeAnalysis],
                       org_analysis: OrganizationAnalysis,
                       original_df: pd.DataFrame,
                       output_path: str) -> bool:
        """종합 Excel 리포트 생성"""
        try:
            wb = Workbook()
            
            # 시트 생성
            self._create_executive_summary(wb, org_analysis)
            self._create_individual_analysis(wb, individual_analyses)
            self._create_competency_analysis(wb, individual_analyses, org_analysis)
            self._create_grade_distribution(wb, org_analysis)
            self._create_department_analysis(wb, org_analysis)
            self._create_position_analysis(wb, org_analysis)
            self._create_recommendations(wb, org_analysis)
            self._create_raw_data(wb, original_df)
            
            # 기본 시트 삭제
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 저장
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Excel generation error: {str(e)}")
            return False
    
    def _create_executive_summary(self, wb: Workbook, org_analysis: OrganizationAnalysis):
        """Executive Summary 시트"""
        ws = wb.create_sheet("Executive Summary")
        
        # 제목
        ws['A1'] = "AIRISS v4.0 조직진단 리포트 - Executive Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # 기본 정보
        ws['A3'] = "총 인원"
        ws['B3'] = org_analysis.total_employees
        ws['D3'] = "분석일시"
        ws['E3'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')
        
        # Executive Summary
        ws['A5'] = "핵심 요약"
        ws['A5'].font = self.header_font
        ws['A5'].fill = self.header_fill
        ws.merge_cells('A5:H5')
        
        # Summary 내용 (줄바꿈 처리)
        summary_lines = org_analysis.executive_summary.split('. ')
        row = 6
        for line in summary_lines:
            if line.strip():
                ws[f'A{row}'] = f"• {line.strip()}"
                ws.merge_cells(f'A{row}:H{row}')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1
        
        # 8대역량 평균
        row += 1
        ws[f'A{row}'] = "8대역량 조직 평균"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        col = 1
        for comp, score in org_analysis.competency_averages.items():
            ws.cell(row=row, column=col, value=comp)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row+1, column=col, value=score)
            col += 1
        
        # 등급 분포 요약
        row += 3
        ws[f'A{row}'] = "등급 분포"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        for grade in ['S', 'A+', 'A', 'B+', 'B', 'C', 'D']:
            count = org_analysis.grade_distribution.get(grade, 0)
            percentage = (count / org_analysis.total_employees * 100) if org_analysis.total_employees > 0 else 0
            ws[f'A{row}'] = f"{grade}등급"
            ws[f'B{row}'] = f"{count}명 ({percentage:.1f}%)"
            
            # 등급별 색상 적용
            if grade in self.grade_colors:
                ws[f'A{row}'].fill = PatternFill(start_color=self.grade_colors[grade], 
                                                 end_color=self.grade_colors[grade], 
                                                 fill_type="solid")
            row += 1
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_individual_analysis(self, wb: Workbook, analyses: List[EmployeeAnalysis]):
        """개인별 상세분석 시트"""
        ws = wb.create_sheet("개인별 상세분석")
        
        # 헤더
        headers = [
            'UID', '이름', '부서', '직급', 'OK등급',
            '성장지향성', '실행력', '혁신', '협업', '고객지향성', '윤리경영', '전문성', '리더십',
            '강점1', '강점2', '강점3',
            '개발영역1', '개발영역2',
            '단기계획', '중장기계획',
            '추천직무', '추천교육',
            'AI 인사이트'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = self.header_font
            ws.cell(row=1, column=col).fill = self.header_fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        
        # 데이터 입력
        row = 2
        for analysis in analyses:
            ws.cell(row=row, column=1, value=analysis.uid)
            ws.cell(row=row, column=2, value=analysis.name)
            ws.cell(row=row, column=3, value=analysis.department)
            ws.cell(row=row, column=4, value=analysis.position)
            ws.cell(row=row, column=5, value=analysis.ok_grade)
            
            # OK등급 색상
            if analysis.ok_grade in self.grade_colors:
                ws.cell(row=row, column=5).fill = PatternFill(
                    start_color=self.grade_colors[analysis.ok_grade],
                    end_color=self.grade_colors[analysis.ok_grade],
                    fill_type="solid"
                )
            
            # 8대역량 점수
            col = 6
            for comp in AIRISSCompetency:
                score = analysis.competency_scores.get(comp.value, 0)
                ws.cell(row=row, column=col, value=score)
                
                # 점수에 따른 색상
                if score >= 4.5:
                    cell_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif score >= 3.5:
                    cell_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif score >= 2.5:
                    cell_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                else:
                    cell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                ws.cell(row=row, column=col).fill = cell_fill
                
                col += 1
            
            # 강점
            for i, strength in enumerate(analysis.strengths[:3]):
                ws.cell(row=row, column=col+i, value=f"{strength['역량']}: {strength['설명']}")
            col += 3
            
            # 개발영역
            for i, dev in enumerate(analysis.development_areas[:2]):
                ws.cell(row=row, column=col+i, value=f"{dev['역량']}: {dev['설명']}")
            col += 2
            
            # 계획
            ws.cell(row=row, column=col, value=analysis.short_term_plan)
            ws.cell(row=row, column=col+1, value=analysis.long_term_plan)
            col += 2
            
            # 추천
            ws.cell(row=row, column=col, value=', '.join(analysis.recommended_positions[:2]))
            ws.cell(row=row, column=col+1, value=', '.join(analysis.recommended_training[:3]))
            col += 2
            
            # AI 인사이트
            ws.cell(row=row, column=col, value=analysis.ai_insights)
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            
            row += 1
        
        # 테두리 적용
        self._apply_borders(ws, 1, 1, row-1, len(headers))
        
        # 필터 적용
        ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{row-1}"
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_competency_analysis(self, wb: Workbook, analyses: List[EmployeeAnalysis], org_analysis: OrganizationAnalysis):
        """8대역량 분석 시트"""
        ws = wb.create_sheet("8대역량 분석")
        
        # 제목
        ws['A1'] = "AIRISS 8대역량 종합 분석"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        # 조직 평균
        ws['A3'] = "조직 전체 평균"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:J3')
        
        row = 4
        ws['A4'] = "역량"
        ws['B4'] = "평균점수"
        ws['C4'] = "최고점수"
        ws['D4'] = "최저점수"
        ws['E4'] = "표준편차"
        ws['F4'] = "S등급 평균"
        ws['G4'] = "A등급 평균"
        ws['H4'] = "B등급 평균"
        ws['I4'] = "C등급 평균"
        ws['J4'] = "Gap(S-C)"
        
        for col in range(1, 11):
            ws.cell(row=4, column=col).font = Font(bold=True)
            ws.cell(row=4, column=col).fill = self.subheader_fill
        
        row = 5
        for comp in AIRISSCompetency:
            scores = [a.competency_scores.get(comp.value, 0) for a in analyses]
            
            ws.cell(row=row, column=1, value=comp.value)
            ws.cell(row=row, column=2, value=round(sum(scores)/len(scores), 2))
            ws.cell(row=row, column=3, value=max(scores))
            ws.cell(row=row, column=4, value=min(scores))
            ws.cell(row=row, column=5, value=round(pd.Series(scores).std(), 2))
            
            # 등급별 평균
            grade_scores = {'S': [], 'A': [], 'B': [], 'C': []}
            for analysis in analyses:
                grade = analysis.ok_grade
                score = analysis.competency_scores.get(comp.value, 0)
                if grade == 'S':
                    grade_scores['S'].append(score)
                elif grade in ['A+', 'A']:
                    grade_scores['A'].append(score)
                elif grade in ['B+', 'B']:
                    grade_scores['B'].append(score)
                else:
                    grade_scores['C'].append(score)
            
            s_avg = round(sum(grade_scores['S'])/len(grade_scores['S']), 2) if grade_scores['S'] else 0
            a_avg = round(sum(grade_scores['A'])/len(grade_scores['A']), 2) if grade_scores['A'] else 0
            b_avg = round(sum(grade_scores['B'])/len(grade_scores['B']), 2) if grade_scores['B'] else 0
            c_avg = round(sum(grade_scores['C'])/len(grade_scores['C']), 2) if grade_scores['C'] else 0
            
            ws.cell(row=row, column=6, value=s_avg)
            ws.cell(row=row, column=7, value=a_avg)
            ws.cell(row=row, column=8, value=b_avg)
            ws.cell(row=row, column=9, value=c_avg)
            ws.cell(row=row, column=10, value=round(s_avg - c_avg, 2))
            
            row += 1
        
        # 차트 추가
        self._add_competency_chart(ws, row-8, row-1)
        
        # 테두리 적용
        self._apply_borders(ws, 1, 4, row-1, 10)
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_grade_distribution(self, wb: Workbook, org_analysis: OrganizationAnalysis):
        """등급 분포 시트"""
        ws = wb.create_sheet("등급 분포")
        
        # 제목
        ws['A1'] = "OK등급 분포 현황"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # 전체 분포
        ws['A3'] = "전체 등급 분포"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:E3')
        
        ws['A4'] = "등급"
        ws['B4'] = "인원"
        ws['C4'] = "비율(%)"
        ws['D4'] = "누적인원"
        ws['E4'] = "누적비율(%)"
        
        for col in range(1, 6):
            ws.cell(row=4, column=col).font = Font(bold=True)
            ws.cell(row=4, column=col).fill = self.subheader_fill
        
        row = 5
        cumulative = 0
        total = org_analysis.total_employees
        
        for grade in ['S', 'A+', 'A', 'B+', 'B', 'C', 'D']:
            count = org_analysis.grade_distribution.get(grade, 0)
            percentage = (count / total * 100) if total > 0 else 0
            cumulative += count
            cum_percentage = (cumulative / total * 100) if total > 0 else 0
            
            ws.cell(row=row, column=1, value=grade)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=round(percentage, 1))
            ws.cell(row=row, column=4, value=cumulative)
            ws.cell(row=row, column=5, value=round(cum_percentage, 1))
            
            # 등급 색상
            if grade in self.grade_colors:
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color=self.grade_colors[grade],
                    end_color=self.grade_colors[grade],
                    fill_type="solid"
                )
            
            row += 1
        
        # 파이 차트 추가
        self._add_grade_pie_chart(ws, org_analysis.grade_distribution)
        
        # 부서별 등급 분포
        row += 2
        ws[f'A{row}'] = "부서별 등급 분포"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:I{row}')
        
        row += 1
        ws[f'A{row}'] = "부서"
        col = 2
        for grade in ['S', 'A+', 'A', 'B+', 'B', 'C', 'D']:
            ws.cell(row=row, column=col, value=grade)
            ws.cell(row=row, column=col).font = Font(bold=True)
            col += 1
        ws.cell(row=row, column=col, value="계")
        ws.cell(row=row, column=col).font = Font(bold=True)
        
        row += 1
        for dept, data in org_analysis.department_analysis.items():
            ws.cell(row=row, column=1, value=dept)
            col = 2
            dept_total = 0
            for grade in ['S', 'A+', 'A', 'B+', 'B', 'C', 'D']:
                count = data['등급분포'].get(grade, 0)
                ws.cell(row=row, column=col, value=count)
                dept_total += count
                col += 1
            ws.cell(row=row, column=col, value=dept_total)
            row += 1
        
        # 테두리 적용
        self._apply_borders(ws, 1, 4, 11, 5)
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_department_analysis(self, wb: Workbook, org_analysis: OrganizationAnalysis):
        """부서별 분석 시트"""
        ws = wb.create_sheet("부서별 분석")
        
        # 제목
        ws['A1'] = "부서별 역량 분석"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        
        # 헤더
        headers = ['부서', '인원'] + [comp.value for comp in AIRISSCompetency] + ['종합평균']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
            ws.cell(row=3, column=col).font = self.header_font
            ws.cell(row=3, column=col).fill = self.header_fill
        
        # 데이터
        row = 4
        for dept, data in org_analysis.department_analysis.items():
            ws.cell(row=row, column=1, value=dept)
            ws.cell(row=row, column=2, value=data['인원'])
            
            col = 3
            total_score = 0
            for comp in AIRISSCompetency:
                score = data['평균역량점수'].get(comp.value, 0)
                ws.cell(row=row, column=col, value=score)
                total_score += score
                
                # 점수별 색상
                if score >= 4.0:
                    cell_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif score < 3.0:
                    cell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                else:
                    cell_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                ws.cell(row=row, column=col).fill = cell_fill
                
                col += 1
            
            # 종합평균
            avg_score = round(total_score / 8, 2)
            ws.cell(row=row, column=col, value=avg_score)
            
            row += 1
        
        # 테두리 적용
        self._apply_borders(ws, 1, 3, row-1, len(headers))
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_position_analysis(self, wb: Workbook, org_analysis: OrganizationAnalysis):
        """직급별 분석 시트"""
        ws = wb.create_sheet("직급별 분석")
        
        # 제목
        ws['A1'] = "직급별 역량 분석"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        
        # 헤더
        headers = ['직급', '인원'] + [comp.value for comp in AIRISSCompetency] + ['종합평균']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
            ws.cell(row=3, column=col).font = self.header_font
            ws.cell(row=3, column=col).fill = self.header_fill
        
        # 직급 순서 정의
        position_order = ['사원', '주임', '대리', '과장', '차장', '부장', '임원']
        
        # 데이터
        row = 4
        for position in position_order:
            if position in org_analysis.position_analysis:
                data = org_analysis.position_analysis[position]
                ws.cell(row=row, column=1, value=position)
                ws.cell(row=row, column=2, value=data['인원'])
                
                col = 3
                total_score = 0
                for comp in AIRISSCompetency:
                    score = data['평균역량점수'].get(comp.value, 0)
                    ws.cell(row=row, column=col, value=score)
                    total_score += score
                    
                    # 점수별 색상
                    if score >= 4.0:
                        cell_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif score < 3.0:
                        cell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    else:
                        cell_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                    ws.cell(row=row, column=col).fill = cell_fill
                    
                    col += 1
                
                # 종합평균
                avg_score = round(total_score / 8, 2)
                ws.cell(row=row, column=col, value=avg_score)
                
                row += 1
        
        # 테두리 적용
        self._apply_borders(ws, 1, 3, row-1, len(headers))
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_recommendations(self, wb: Workbook, org_analysis: OrganizationAnalysis):
        """HR 전략 제언 시트"""
        ws = wb.create_sheet("HR 전략 제언")
        
        # 제목
        ws['A1'] = "HR 전략 제언 및 실행 계획"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # 전략적 제언
        ws['A3'] = "전략적 제언"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:D3')
        
        row = 4
        for i, rec in enumerate(org_analysis.strategic_recommendations, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = rec
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # 리스크
        row += 1
        ws[f'A{row}'] = "주요 리스크"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for i, risk in enumerate(org_analysis.risks, 1):
            ws[f'A{row}'] = f"⚠️ {i}."
            ws[f'B{row}'] = risk
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # 기회요인
        row += 1
        ws[f'A{row}'] = "기회요인"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for i, opp in enumerate(org_analysis.opportunities, 1):
            ws[f'A{row}'] = f"✓ {i}."
            ws[f'B{row}'] = opp
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # 실행 로드맵
        row += 2
        ws[f'A{row}'] = "실행 로드맵"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "구분"
        ws[f'B{row}'] = "단기 (3개월)"
        ws[f'C{row}'] = "중기 (6개월)"
        ws[f'D{row}'] = "장기 (1년)"
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = self.subheader_fill
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
    
    def _create_raw_data(self, wb: Workbook, df: pd.DataFrame):
        """원본 데이터 시트"""
        ws = wb.create_sheet("원본 데이터")
        
        # 데이터프레임을 시트에 쓰기
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 헤더 스타일
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # 필터 적용
        ws.auto_filter.ref = ws.dimensions
        
        # 열 너비 자동 조정
        self._adjust_column_widths(ws)
    
    def _add_competency_chart(self, ws, start_row: int, end_row: int):
        """역량 차트 추가"""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "8대역량 조직 평균"
        chart.y_axis.title = '평균 점수'
        chart.x_axis.title = '역량'
        
        data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row, max_col=2)
        cats = Reference(ws, min_col=1, min_row=start_row+1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        
        ws.add_chart(chart, "L4")
    
    def _add_grade_pie_chart(self, ws, grade_dist: Dict):
        """등급 분포 파이 차트"""
        pie = PieChart()
        labels = []
        data_values = []
        
        for grade in ['S', 'A+', 'A', 'B+', 'B', 'C', 'D']:
            if grade in grade_dist and grade_dist[grade] > 0:
                labels.append(grade)
                data_values.append(grade_dist[grade])
        
        # 차트 데이터를 시트에 임시로 쓰기
        start_row = 15
        ws['H14'] = "Chart Data"
        for i, (label, value) in enumerate(zip(labels, data_values)):
            ws.cell(row=start_row+i, column=8, value=label)
            ws.cell(row=start_row+i, column=9, value=value)
        
        data = Reference(ws, min_col=9, min_row=start_row, max_row=start_row+len(labels)-1)
        categories = Reference(ws, min_col=8, min_row=start_row, max_row=start_row+len(labels)-1)
        
        pie.add_data(data)
        pie.set_categories(categories)
        pie.title = "등급 분포"
        
        ws.add_chart(pie, "G5")
    
    def _apply_borders(self, ws, start_row: int, start_col: int, end_row: int, end_col: int):
        """테두리 적용"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.border
    
    def _adjust_column_widths(self, ws):
        """열 너비 자동 조정"""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                if adjusted_width > 0:
                    ws.column_dimensions[column_letter].width = adjusted_width